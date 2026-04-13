# exporter.py
# Exports dashboard data to a formatted Excel report


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import pandas as pd
from data_generator import generate_weekly_data, generate_pdp_scores
from metrics import summary_kpis


BRAND_BLUE  = "185FA5"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_BLUE)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D3D1C7"),
    right =Side(style="thin", color="D3D1C7"),
)


def style_header_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if r_idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    # auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 32)


def export_report(path: str = "coty_digital_shelf_report.xlsx"):
    """
    Produces a three-sheet Excel report:
      1. KPI Summary  — headline metrics with delta
      2. Weekly Data  — full campaign time series
      3. PDP Tracker  — content completeness by SKU
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: KPI summary ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "KPI Summary"
    df_weekly = generate_weekly_data()
    kpis = summary_kpis(df_weekly)

    ws1["A1"] = "Coty Digital Shelf — Q1 2025 Performance Summary"
    ws1["A1"].font = Font(bold=True, size=13, color=BRAND_BLUE)
    ws1.merge_cells("A1:D1")

    headers = ["KPI", "Value", "Delta vs prior period", "Status"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=3, column=c, value=h)
    style_header_row(ws1, 3, 4)

    for i, (kpi, data) in enumerate(kpis.items(), 4):
        delta = data["delta"]
        status = "✓ On track" if delta > 0 else "⚠ Review"
        ws1.cell(row=i, column=1, value=kpi.replace("_", " ").title())
        ws1.cell(row=i, column=2, value=data["value"])
        ws1.cell(row=i, column=3, value=f"{delta:+.1f}%")
        ws1.cell(row=i, column=4, value=status)

    # ── Sheet 2: Weekly campaign data ────────────────────────────
    ws2 = wb.create_sheet("Weekly Data")
    write_dataframe(ws2, df_weekly)

    # ── Sheet 3: PDP tracker ─────────────────────────────────────
    ws3 = wb.create_sheet("PDP Tracker")
    write_dataframe(ws3, generate_pdp_scores())

    wb.save(path)
    print(ff"Report saved → {path}")


if __name__ == "__main__":
    export_report()
